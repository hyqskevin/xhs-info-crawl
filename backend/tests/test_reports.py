from datetime import datetime, timezone
from io import BytesIO
from time import perf_counter

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.security import create_access_token
from app.models.activity import Activity
from app.services.report import format_activity_markdown, generate_markdown, generate_xlsx


@pytest.fixture
def headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token({'sub': 'admin', 'role': 'admin'})}"}


def activity(index: int, city: str = "shanghai", kind: str = "演出", status: str = "APPROVED") -> Activity:
    return Activity(name=f"活动{index}", city_code=city, start_time=datetime(2025, 7, 20, 18, tzinfo=timezone.utc), end_time=datetime(2025, 7, 20, 22, tzinfo=timezone.utc), location="徐汇滨江", price="免费", type=kind, source_url=f"https://www.xiaohongshu.com/a/{index}", summary=f"活动{index}简介", status=status)


def test_report_groups_by_city_and_type_and_excludes_non_approved() -> None:
    items = [activity(1), activity(2, kind="展览"), activity(3, "beijing"), activity(4, status="NEEDS_REVIEW")]
    report = generate_markdown("2025-W29", ["shanghai", "beijing"], items)
    assert "## 上海" in report and "## 北京" in report
    assert "### 演出" in report and "### 展览" in report
    assert "活动4" not in report


def test_activity_markdown_format_and_empty_report() -> None:
    text = format_activity_markdown(activity(1))
    assert "#### 活动1" in text
    assert "**时间**：2025-07-20 18:00 - 22:00" in text
    assert "**地点**：徐汇滨江" in text and "**费用**：免费" in text
    assert "[小红书笔记](https://www.xiaohongshu.com/a/1)" in text
    assert "本周暂无活动" in generate_markdown("2025-W29", ["shanghai"], [])


def test_report_generation_performance() -> None:
    items = [activity(i, kind=["演出", "展览", "市集", "沙龙"][i % 4]) for i in range(500)]
    started = perf_counter()
    generate_markdown("2025-W29", ["shanghai"], items)
    assert perf_counter() - started <= 30


def test_xlsx_contains_same_approved_activities() -> None:
    content = generate_xlsx([activity(1), activity(2, status="IGNORED")])
    workbook = load_workbook(BytesIO(content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[1][0] == "活动1"
    assert len(rows) == 2


def test_generate_persists_and_regenerates_single_report(client: TestClient, db_session: Session, headers: dict[str, str]) -> None:
    db_session.add(activity(1))
    db_session.commit()
    payload = {"week": "2025-W29", "cities": ["shanghai"]}
    first = client.post("/api/v1/reports/generate", json=payload, headers=headers)
    second = client.post("/api/v1/reports/generate", json=payload, headers=headers)
    assert first.status_code == 200 and second.status_code == 200
    assert first.json()["data"]["id"] == second.json()["data"]["id"]
    assert second.json()["data"]["activity_count"] == 1


def test_report_generation_requires_exactly_one_city(client: TestClient, headers: dict[str, str]) -> None:
    assert client.post("/api/v1/reports/generate", json={"week": "2026-W29", "cities": []}, headers=headers).status_code == 422
    assert client.post("/api/v1/reports/generate", json={"week": "2026-W29", "cities": ["nb", "shanghai"]}, headers=headers).status_code == 422


def test_download_report_returns_markdown_and_excel(client: TestClient, db_session: Session, headers: dict[str, str]) -> None:
    db_session.add(activity(1))
    db_session.commit()
    report_id = client.post("/api/v1/reports/generate", json={"week": "2025-W29", "cities": ["shanghai"]}, headers=headers).json()["data"]["id"]
    md = client.get(f"/api/v1/reports/{report_id}/download?format=md", headers=headers)
    xlsx = client.get(f"/api/v1/reports/{report_id}/download?format=xlsx", headers=headers)
    assert md.status_code == 200 and "text/markdown" in md.headers["content-type"]
    assert xlsx.status_code == 200 and "spreadsheetml.sheet" in xlsx.headers["content-type"]
    assert "2025-W29" in md.headers["content-disposition"] and "2025-W29" in xlsx.headers["content-disposition"]
