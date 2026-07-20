from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.models.blogger_city import BloggerCity
from app.models.config import Blogger, City
from app.core.security import create_access_token
from app.services.blogger_import import BloggerImportError, generate_blogger_template, import_bloggers


HEADERS = ["博主名称", "小红书用户ID", "主页地址", "关联城市", "启用"]


def auth_headers():
    return {"Authorization": f"Bearer {create_access_token({'sub': 'admin', 'role': 'admin'})}"}


def xlsx_bytes(*rows) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def add_cities(db_session):
    db_session.add_all([
        City(name="宁波", code="nb", enabled=True),
        City(name="上海", code="sh", enabled=True),
    ])
    db_session.commit()


def test_template_contains_user_facing_columns():
    sheet = load_workbook(BytesIO(generate_blogger_template()), read_only=True).active
    assert list(next(sheet.iter_rows(values_only=True))) == HEADERS


def test_xlsx_import_creates_bloggers_and_splits_city_names(db_session):
    add_cities(db_session)

    result = import_bloggers(db_session, xlsx_bytes(
        ("博主甲", "uid-a", "https://xhs/a", "宁波、上海", "是"),
        ("博主乙", "", "", "宁波", "否"),
    ), "bloggers.xlsx")

    assert result == {"created": 2, "updated": 0, "total": 2}
    bloggers = list(db_session.scalars(select(Blogger).order_by(Blogger.username)))
    assert [(item.username, item.enabled) for item in bloggers] == [("博主乙", False), ("博主甲", True)]
    bindings = list(db_session.execute(select(Blogger.username, BloggerCity.city_code).join(BloggerCity).order_by(Blogger.username, BloggerCity.city_code)))
    assert bindings == [("博主乙", "nb"), ("博主甲", "nb"), ("博主甲", "sh")]


def test_csv_reimport_is_idempotent_and_replaces_city_bindings(db_session):
    add_cities(db_session)
    first = "博主名称,小红书用户ID,主页地址,关联城市,启用\n博主甲,uid-a,https://xhs/a,宁波,true\n".encode()
    second = "博主名称,小红书用户ID,主页地址,关联城市,启用\n博主甲更新,uid-a,https://xhs/a,上海,1\n".encode()

    assert import_bloggers(db_session, first, "list.csv")["created"] == 1
    assert import_bloggers(db_session, second, "list.csv") == {"created": 0, "updated": 1, "total": 1}
    assert db_session.scalar(select(func.count()).select_from(Blogger)) == 1
    blogger = db_session.scalar(select(Blogger))
    assert blogger.username == "博主甲更新"
    assert list(db_session.scalars(select(BloggerCity.city_code).where(BloggerCity.blogger_id == blogger.id))) == ["sh"]


def test_validation_error_has_row_and_writes_nothing(db_session):
    add_cities(db_session)
    content = xlsx_bytes(
        ("合法博主", "uid-a", "", "宁波", "是"),
        ("错误博主", "uid-b", "", "不存在城市", "是"),
    )

    with pytest.raises(BloggerImportError, match="第3行.*不存在城市"):
        import_bloggers(db_session, content, "list.xlsx")

    assert db_session.scalar(select(func.count()).select_from(Blogger)) == 0


def test_duplicate_identity_in_file_is_rejected(db_session):
    add_cities(db_session)
    content = xlsx_bytes(
        ("博主甲", "uid-a", "", "宁波", "是"),
        ("博主甲副本", "uid-a", "", "上海", "是"),
    )

    with pytest.raises(BloggerImportError, match="第3行.*重复"):
        import_bloggers(db_session, content, "list.xlsx")


@pytest.mark.parametrize("filename", ["list.xls", "list.json"])
def test_unsupported_file_type_is_rejected(db_session, filename):
    with pytest.raises(BloggerImportError, match="仅支持"):
        import_bloggers(db_session, b"bad", filename)


def test_import_template_api_downloads_xlsx(client):
    response = client.get("/api/v1/settings/bloggers/import-template", headers=auth_headers())

    assert response.status_code == 200
    assert "spreadsheetml.sheet" in response.headers["content-type"]
    assert "blogger-import-template.xlsx" in response.headers["content-disposition"]
    assert list(next(load_workbook(BytesIO(response.content), read_only=True).active.iter_rows(values_only=True))) == HEADERS


def test_import_api_accepts_raw_xlsx_and_returns_counts(client, db_session):
    add_cities(db_session)

    response = client.post(
        "/api/v1/settings/bloggers/import",
        params={"filename": "list.xlsx"},
        content=xlsx_bytes(("博主甲", "uid-a", "", "宁波", "是")),
        headers={**auth_headers(), "Content-Type": "application/octet-stream"},
    )

    assert response.status_code == 201
    assert response.json()["data"] == {"created": 1, "updated": 0, "total": 1}


def test_import_api_returns_row_error_without_partial_write(client, db_session):
    add_cities(db_session)

    response = client.post(
        "/api/v1/settings/bloggers/import",
        params={"filename": "list.xlsx"},
        content=xlsx_bytes(("博主甲", "uid-a", "", "未知城市", "是")),
        headers=auth_headers(),
    )

    assert response.status_code == 422
    assert "第2行" in response.json()["message"]
    assert db_session.scalar(select(func.count()).select_from(Blogger)) == 0


def test_import_api_rejects_files_over_two_mib(client):
    response = client.post(
        "/api/v1/settings/bloggers/import",
        params={"filename": "list.csv"},
        content=b"x" * (2 * 1024 * 1024 + 1),
        headers=auth_headers(),
    )

    assert response.status_code == 413


def test_import_api_requires_admin(client):
    assert client.get("/api/v1/settings/bloggers/import-template").status_code == 401
