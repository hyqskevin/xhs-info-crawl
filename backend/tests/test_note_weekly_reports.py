from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.api.v1.reports import select_notes
from app.models.activity import Activity
from app.models.note import Note, NoteImage
from app.services.report import generate_note_markdown, generate_note_xlsx


def test_weekly_report_selects_approved_posts_by_publish_time_and_keeps_all_children(db_session) -> None:
    note = Note(
        task_id=1,
        platform_note_id="weekly-post",
        title="本周发布的活动合集",
        content="推文正文",
        source_url="https://www.xiaohongshu.com/explore/weekly-post",
        city_code="nb",
        status="PROCESSED",
        review_status="APPROVED",
        published_at=datetime(2026, 7, 15, 12, tzinfo=timezone.utc),
        raw_data={},
    )
    db_session.add(note)
    db_session.flush()
    db_session.add_all([
        Activity(note_id=note.id, name="本周活动", city_code="nb", type="展览", status="RAW", start_time=datetime(2026, 7, 17, tzinfo=timezone.utc)),
        Activity(note_id=note.id, name="下月活动", city_code="nb", type="演出", status="NEEDS_REVIEW", start_time=datetime(2026, 8, 17, tzinfo=timezone.utc)),
        NoteImage(note_id=note.id, storage_key="images/post.jpg", original_url="https://img.example/post.jpg", ocr_text="OCR文字", ocr_status="success"),
    ])
    db_session.commit()

    entries = select_notes(db_session, ["nb"], "2026-W29")

    assert len(entries) == 1
    selected_note, activities, images = entries[0]
    assert selected_note.id == note.id
    assert [item.name for item in activities] == ["本周活动", "下月活动"]

    markdown = generate_note_markdown("2026-W29", ["nb"], entries)
    assert note.source_url in markdown
    assert "https://img.example/post.jpg" in markdown
    assert "本周活动" in markdown and "下月活动" in markdown

    workbook = load_workbook(BytesIO(generate_note_xlsx(entries)))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert len(rows) == 2
    assert rows[1][0] == note.title
    assert "本周活动" in rows[1][-1] and "下月活动" in rows[1][-1]
