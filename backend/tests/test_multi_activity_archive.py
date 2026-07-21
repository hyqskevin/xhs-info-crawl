from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.models.activity import Activity
from app.models.note import Note, NoteImage
from app.services.archive import archive_task_result, resolve_storage_path
from app.services.extraction import extract_activities


def test_extract_activities_returns_every_llm_activity_with_source_images():
    def llm(_: str):
        return {
            "activities": [
                {"name": "滨江音乐会", "start_time": "2026-07-18T19:00:00", "location": "徐汇滨江", "price": "免费", "type": "演出", "summary": "露天演出", "confidence": 0.92, "source_image_indexes": [1, 2]},
                {"name": "周末手作市集", "start_time": "2026-07-19T10:00:00", "location": "凯德晶萃广场", "price": "免费", "type": "市集", "summary": "手作摊位", "confidence": "high", "source_image_indexes": [3]},
            ]
        }

    result = extract_activities("[IMAGE 1]\n音乐会\n[IMAGE 3]\n市集", datetime(2026, 7, 16), llm)
    assert [item["name"] for item in result] == ["滨江音乐会", "周末手作市集"]
    assert result[0]["source_image_indexes"] == [1, 2]
    assert result[1]["confidence"] == 0.9
    assert "status" not in result[0]


def test_llm_partial_dates_are_normalized_without_crashing():
    result = extract_activities("活动", datetime(2026, 7, 17), lambda _: {"activities": [
        {"name": "春日市集", "start_time": "4/5", "end_time": "非法日期", "location": "月湖公园", "source_image_indexes": [1]},
        {"name": "夏日市集", "start_time": "8/5", "location": "月湖公园", "source_image_indexes": [2]},
        {"name": "无效日期活动", "start_time": "2/30", "location": "文化广场", "source_image_indexes": []},
    ]})

    # 新规则：无 60 天上限；4/5 / 2/30 解析后保留返回值，2/30 因非法日期保留字符串无效特征
    assert all("status" not in item for item in result)
    # 8/5 对比 now=2026-07-17 是未来月份，保留本年 -> 2026-08-05
    assert result[1]["start_time"] == "2026-08-05T00:00:00"


def test_llm_date_with_ambiguous_chinese_time_is_rejected_safely():
    result = extract_activities("活动", datetime(2026, 7, 17), lambda _: {"activities": [
        {"name": "晚间活动", "start_time": "2026-07-18T晚间", "location": "文化广场", "source_image_indexes": []},
    ]})[0]

    assert result["start_time"] is None
    assert "status" not in result


def test_archive_places_source_images_markdown_and_xlsx_under_date_task_folder(tmp_path: Path):
    note = Note(id=7, task_id=9, platform_note_id="note-7", title="上海周末合集", content="原文正文", source_url="https://www.xiaohongshu.com/explore/note-7", city_code="shanghai", status="OCR_DONE", raw_data={})
    image_file = tmp_path / "download.jpg"
    image_file.write_bytes(b"image")
    images = [NoteImage(id=1, note_id=7, storage_key="", ocr_text="滨江音乐会 7月18日", ocr_status="success")]
    activities = [Activity(id=3, note_id=7, name="滨江音乐会", city_code="shanghai", start_time=datetime(2026, 7, 18, 19, tzinfo=timezone.utc), location="徐汇滨江", price="免费", type="演出", source_url=note.source_url, summary="露天演出", source_image_indexes=[1])]

    folder = archive_task_result(tmp_path / "archive", datetime(2026, 7, 16, tzinfo=timezone.utc), 9, note, [(image_file, images[0])], activities)

    assert folder == tmp_path / "archive" / "2026-07-16" / "task-9"
    assert (folder / "images" / "note-7_01.jpg").exists()
    source_markdown=(folder / "source.md").read_text(encoding="utf-8")
    activities_markdown=(folder / "activities.md").read_text(encoding="utf-8")
    assert note.source_url in source_markdown
    assert "![图片 1](images/note-7_01.jpg)" in source_markdown
    assert "滨江音乐会" in activities_markdown
    assert "[来源图片 1](images/note-7_01.jpg)" in activities_markdown
    rows = list(load_workbook(folder / "activities.xlsx", read_only=True).active.iter_rows(values_only=True))
    assert len(rows) == 2
    assert rows[1][0] == "滨江音乐会"
    # xlsx 列：name | start | end | location | price | type | images | source_url | summary
    assert rows[1][7] == note.source_url
    assert rows[1][8] == "露天演出"


def test_resolve_storage_path_supports_legacy_image_relative_keys(tmp_path: Path):
    image=(tmp_path/'images'/'note'/'a.jpg'); image.parent.mkdir(parents=True); image.write_bytes(b'image')
    assert resolve_storage_path(tmp_path,tmp_path/'images','note/a.jpg')==image
