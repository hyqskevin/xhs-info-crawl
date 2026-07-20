import csv
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.blogger_city import BloggerCity
from app.models.config import Blogger, City


HEADERS = ["博主名称", "小红书用户ID", "主页地址", "关联城市", "启用"]
MAX_ROWS = 500
CITY_SEPARATOR = re.compile(r"[、,，;；]+")


class BloggerImportError(ValueError):
    pass


@dataclass(frozen=True)
class BloggerImportRow:
    row_number: int
    username: str
    platform_user_id: str | None
    profile_url: str | None
    city_names: list[str]
    enabled: bool


def generate_blogger_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "博主白名单"
    sheet.append(HEADERS)
    sheet.append(["示例博主", "可选", "https://www.xiaohongshu.com/user/profile/可选", "宁波、上海", "是"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _enabled(value: str, row_number: int) -> bool:
    normalized = value.strip().lower()
    if normalized in {"", "是", "true", "1"}:
        return True
    if normalized in {"否", "false", "0"}:
        return False
    raise BloggerImportError(f"第{row_number}行：启用值无效")


def _rows_from_file(content: bytes, filename: str) -> list[list[object]]:
    suffix = Path(filename).suffix.lower()
    try:
        if suffix == ".xlsx":
            sheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        if suffix == ".csv":
            return [list(row) for row in csv.reader(StringIO(content.decode("utf-8-sig")))]
    except Exception as exc:
        raise BloggerImportError(f"文件解析失败：{exc}") from exc
    raise BloggerImportError("仅支持 .xlsx 或 UTF-8 .csv 文件")


def parse_blogger_import(content: bytes, filename: str) -> list[BloggerImportRow]:
    raw_rows = _rows_from_file(content, filename)
    if not raw_rows or [_text(value) for value in raw_rows[0][:5]] != HEADERS:
        raise BloggerImportError("第1行：表头必须与模板一致")
    result: list[BloggerImportRow] = []
    for row_number, raw in enumerate(raw_rows[1:], 2):
        values = list(raw[:5]) + [None] * max(0, 5 - len(raw))
        username, platform_user_id, profile_url, cities, enabled = (_text(value) for value in values[:5])
        if not any((username, platform_user_id, profile_url, cities, enabled)):
            continue
        if not username:
            raise BloggerImportError(f"第{row_number}行：博主名称不能为空")
        if len(username) > 128:
            raise BloggerImportError(f"第{row_number}行：博主名称超过128字符")
        city_names = list(dict.fromkeys(name.strip() for name in CITY_SEPARATOR.split(cities) if name.strip()))
        if not city_names:
            raise BloggerImportError(f"第{row_number}行：关联城市不能为空")
        result.append(BloggerImportRow(
            row_number=row_number,
            username=username,
            platform_user_id=platform_user_id or None,
            profile_url=profile_url or None,
            city_names=city_names,
            enabled=_enabled(enabled, row_number),
        ))
    if not result:
        raise BloggerImportError("文件没有有效数据行")
    if len(result) > MAX_ROWS:
        raise BloggerImportError(f"有效数据行不能超过{MAX_ROWS}行")
    return result


def import_bloggers(db: Session, content: bytes, filename: str) -> dict[str, int]:
    rows = parse_blogger_import(content, filename)
    cities = list(db.scalars(select(City)).all())
    city_by_name: dict[str, City] = {}
    duplicate_city_names: set[str] = set()
    for city in cities:
        if city.name in city_by_name:
            duplicate_city_names.add(city.name)
        city_by_name[city.name] = city

    bloggers = list(db.scalars(select(Blogger)).all())
    by_user_id = {item.platform_user_id: item for item in bloggers if item.platform_user_id}
    by_profile = {(item.profile_url or "").strip(): item for item in bloggers if (item.profile_url or "").strip()}
    by_username = {item.username.strip(): item for item in bloggers}
    seen_user_ids: set[str] = set()
    seen_profiles: set[str] = set()
    seen_usernames: set[str] = set()
    seen_targets: set[int] = set()
    plan: list[tuple[BloggerImportRow, Blogger | None, list[str]]] = []

    for row in rows:
        for city_name in row.city_names:
            if city_name in duplicate_city_names:
                raise BloggerImportError(f"第{row.row_number}行：城市名称存在重复配置：{city_name}")
            if city_name not in city_by_name:
                raise BloggerImportError(f"第{row.row_number}行：不存在城市：{city_name}")
        if row.platform_user_id and row.platform_user_id in seen_user_ids:
            raise BloggerImportError(f"第{row.row_number}行：小红书用户ID重复")
        if row.profile_url and row.profile_url in seen_profiles:
            raise BloggerImportError(f"第{row.row_number}行：主页地址重复")
        if row.username in seen_usernames:
            raise BloggerImportError(f"第{row.row_number}行：博主名称重复")
        if row.platform_user_id:
            seen_user_ids.add(row.platform_user_id)
        if row.profile_url:
            seen_profiles.add(row.profile_url)
        seen_usernames.add(row.username)

        matches = {
            item.id: item
            for item in (
                by_user_id.get(row.platform_user_id) if row.platform_user_id else None,
                by_profile.get(row.profile_url) if row.profile_url else None,
                by_username.get(row.username),
            )
            if item is not None
        }
        if len(matches) > 1:
            raise BloggerImportError(f"第{row.row_number}行：身份字段匹配到不同博主")
        target = next(iter(matches.values()), None)
        if target and target.id in seen_targets:
            raise BloggerImportError(f"第{row.row_number}行：与前一行指向同一博主，重复")
        if target:
            seen_targets.add(target.id)
        plan.append((row, target, [city_by_name[name].code for name in row.city_names]))

    created = 0
    updated = 0
    try:
        for row, target, city_codes in plan:
            if target is None:
                target = Blogger(username=row.username, platform_user_id=row.platform_user_id, profile_url=row.profile_url, enabled=row.enabled)
                db.add(target)
                db.flush()
                created += 1
            else:
                target.username = row.username
                if row.platform_user_id:
                    target.platform_user_id = row.platform_user_id
                if row.profile_url:
                    target.profile_url = row.profile_url
                target.enabled = row.enabled
                updated += 1
            db.execute(delete(BloggerCity).where(BloggerCity.blogger_id == target.id))
            db.add_all(BloggerCity(blogger_id=target.id, city_code=code, enabled=True) for code in city_codes)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"created": created, "updated": updated, "total": len(plan)}
