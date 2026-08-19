import logging
import re
import zipfile
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage

from app.core.config import get_settings
from app.models.activity import Activity
from app.models.note import Note, NoteImage

logger = logging.getLogger(__name__)


CITY_NAMES = {"shanghai": "上海", "beijing": "北京"}

# 匹配周报生成的内联图片行，如：![图片 1](/api/v1/reports/image/xxx.jpg)
_IMAGE_LINE_RE = re.compile(r"^\s*!\[图片 \d+\]\([^)]*\)\s*$")


def _image_markdown_src(image: NoteImage) -> str | None:
    """返回图片在预览里的可访问 URL：优先本地 storage_key，回退 original_url。"""
    if image.storage_key:
        prefix = get_settings().api_v1_prefix
        return f"{prefix}/reports/image/{quote(image.storage_key)}"
    if image.original_url:
        return image.original_url
    return None


def strip_report_images(markdown: str) -> str:
    """去掉周报正文里生成的内联图片行（用于 md 下载，不输出图片地址）。"""
    return "\n".join(line for line in markdown.splitlines() if not _IMAGE_LINE_RE.match(line))


def format_activity_markdown(activity: Activity) -> str:
    start = activity.start_time.strftime("%Y-%m-%d %H:%M") if activity.start_time else "待确认"
    end = activity.end_time.strftime("%H:%M") if activity.end_time else ""
    time_text = f"{start} - {end}" if end else start
    return f"#### {activity.name}\n- **时间**：{time_text}\n- **地点**：{activity.location}\n- **费用**：{activity.price}\n- **来源**：[小红书笔记]({activity.source_url})\n- **简介**：{activity.summary}\n"


NoteReportEntry = tuple[Note, list[Activity], list[NoteImage]]


def _activity_lines(activities: list[Activity]) -> str:
    return "\n".join(
        f"{item.name} | {item.start_time.isoformat() if item.start_time else '时间待确认'} | {item.location} | {item.price} | {item.summary}"
        for item in activities
    )


def generate_note_markdown(week: str, cities: list[str], entries: list[NoteReportEntry]) -> str:
    lines = [f"# 本周推文周报（{week}）", "", f"城市：{'、'.join(CITY_NAMES.get(city, city) for city in cities)}", ""]
    for note, activities, images in entries:
        published = note.published_at.isoformat() if note.published_at else f"{note.created_at.isoformat()}（发布时间待确认）"
        image_lines = [f"![图片 {index}]({src})" for index, src in enumerate(
            (s for s in (_image_markdown_src(image) for image in images) if s), 1
        )]
        ocr = "\n\n".join(image.ocr_text for image in images if image.ocr_text)
        lines.extend([
            f"## {note.title}", "",
            f"- 发布时间：{published}",
            f"- 原文链接：{note.source_url}", "",
        ])
        for line in image_lines:
            lines.extend([line, ""])
        lines.extend([
            "### 推文正文", "", note.content or "无", "",
            "### 图片 OCR", "", ocr or "无", "",
            f"### 识别活动（{len(activities)}）", "",
        ])
        if activities:
            for item in activities:
                lines.extend([format_activity_markdown(item), ""])
        else:
            lines.extend(["未识别到活动", ""])
    return "\n".join(lines)


def generate_note_xlsx(entries: list[NoteReportEntry]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "本周推文"
    sheet.append(["推文标题", "发布时间", "城市", "原文链接", "正文", "OCR", "活动数", "活动详情"])
    for note, activities, images in entries:
        published = note.published_at.isoformat() if note.published_at else f"{note.created_at.isoformat()}（待确认）"
        ocr = "\n".join(image.ocr_text for image in images if image.ocr_text)
        sheet.append([note.title, published, CITY_NAMES.get(note.city_code, note.city_code), note.source_url, note.content, ocr, len(activities), _activity_lines(activities)])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# 周报压缩包（zip）+ xlsx 封面图（2026-08-18 用户反馈）
# 关联 spec: docs/superpowers/specs/2026-08-18-weekly-report-images-and-zip-design.md
# ---------------------------------------------------------------------------


def _sanitize_zip_filename(name: str, max_length: int = 80) -> str:
    """清洗 zip 内的 md 文件名，去掉路径分隔符与控制字符，长度截断到 max_length。"""
    cleaned = re.sub(r"[^\w\u4e00-\u9fff-]", "_", name or "")
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return (cleaned or "report")[:max_length]


def _resolve_local_image(image: NoteImage, data_root: Path) -> Path | None:
    """解析图片本地路径，校验在 data_root 内且存在；不满足返回 None 并记 WARNING。"""
    if not image.storage_key:
        return None
    try:
        target = (data_root / image.storage_key).resolve()
    except (OSError, ValueError):
        logger.warning("storage_key 解析失败 note_id=%s storage_key=%s", image.note_id, image.storage_key)
        return None
    if not target.is_relative_to(data_root) or not target.is_file():
        logger.warning("图片不在 data_dir 内或不存在 note_id=%s storage_key=%s", image.note_id, image.storage_key)
        return None
    return target


def _build_md_for_zip(
    week: str, cities: list[str], entries: list[NoteReportEntry], data_root: Path
) -> tuple[str, list[tuple[Path, str]]]:
    """生成 zip 内 md（图片引用为相对路径 `images/note_<id>/<filename>`）并返回打包文件清单。

    返回 (markdown_text, [(本地绝对路径, zip 内归档路径), ...])。
    """
    lines: list[str] = [f"# 本周推文周报（{week}）", "", f"城市：{'、'.join(CITY_NAMES.get(city, city) for city in cities)}", ""]
    files: list[tuple[Path, str]] = []
    for note, activities, images in entries:
        published = note.published_at.isoformat() if note.published_at else f"{note.created_at.isoformat()}（发布时间待确认）"
        lines.extend([
            f"## {note.title}", "",
            f"- 发布时间：{published}",
            f"- 原文链接：{note.source_url}", "",
        ])
        local_image_index = 0
        for image in images:
            local_path = _resolve_local_image(image, data_root)
            if local_path is None:
                continue
            local_image_index += 1
            archive_name = f"note_{note.id}/{local_path.name}"
            lines.extend([f"![图片 {local_image_index}](images/{archive_name})", ""])
            files.append((local_path, f"images/{archive_name}"))
        lines.extend([
            "### 推文正文", "", note.content or "无", "",
            "### 图片 OCR", "",
            "\n\n".join(image.ocr_text for image in images if image.ocr_text) or "无", "",
            f"### 识别活动（{len(activities)}）", "",
        ])
        if activities:
            for item in activities:
                lines.extend([format_activity_markdown(item), ""])
        else:
            lines.extend(["未识别到活动", ""])
    return "\n".join(lines), files


def build_report_zip(report_name: str, week: str, cities: list[str], entries: list[NoteReportEntry]) -> bytes:
    """打包周报压缩包：<sanitized_name>.md + images/note_<id>/<filename>。

    zip 文件名 UTF-8 编码（依赖 Python 3.10+ 的 zipfile 默认 ZIP_UTF8 行为）；
    md 内图片引用使用相对路径，方便用户解压后用任意 markdown 阅读器查看。
    """
    data_root = get_settings().data_dir.resolve()
    md_text, files = _build_md_for_zip(week, cities, entries, data_root)
    safe_name = _sanitize_zip_filename(report_name or week)
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{safe_name}.md", md_text)
        for local_path, archive_name in files:
            zf.write(local_path, archive_name)
    return buffer.getvalue()


def _maybe_resize_cover_image(local_path: Path, max_bytes: int = 2 * 1024 * 1024) -> bytes:
    """若图片 > max_bytes 且 PIL 可用则缩放到最长边 ≤ 1024px；否则原样返回 bytes。

    缺 PIL → 记录 WARNING 后直接返回原文件 bytes（不阻塞导出）。
    """
    raw = local_path.read_bytes()
    if len(raw) <= max_bytes:
        return raw
    try:
        from io import BytesIO as _BIO
        from PIL import Image as PILImage  # type: ignore[import-not-found]

        with PILImage.open(_BIO(raw)) as img:
            img = img.convert("RGB")
            longest = max(img.size)
            if longest > 1024:
                scale = 1024 / longest
                new_size = (max(1, int(img.size[0] * scale)), max(1, int(img.size[1] * scale)))
                img = img.resize(new_size, PILImage.LANCZOS)
            out = _BIO()
            img.save(out, format="JPEG", quality=85, optimize=True)
            return out.getvalue()
    except ImportError:
        logger.warning("Pillow 未安装，封面图 %s 大小 %d bytes 仍按原图嵌入", local_path, len(raw))
        return raw
    except Exception as exc:  # pragma: no cover - PIL 解码失败等异常
        logger.warning("封面图缩放失败 %s: %s", local_path, exc)
        return raw


def generate_note_xlsx_with_cover(entries: list[NoteReportEntry], data_root: Path) -> bytes:
    """xlsx 末尾新增「封面图」列，嵌入每条 note 第一张本地图片。

    缺图时该列写 "—"；图片 > 2MB 自动 PIL 缩放到最长边 ≤ 1024px；
    行高固定 96px（适配大多数缩略图）。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "本周推文"
    sheet.append(["推文标题", "发布时间", "城市", "原文链接", "正文", "OCR", "活动数", "活动详情", "封面图"])
    for row_index, (note, activities, images) in enumerate(entries, start=2):
        published = note.published_at.isoformat() if note.published_at else f"{note.created_at.isoformat()}（待确认）"
        ocr = "\n".join(image.ocr_text for image in images if image.ocr_text)
        sheet.append([note.title, published, CITY_NAMES.get(note.city_code, note.city_code), note.source_url, note.content, ocr, len(activities), _activity_lines(activities), "—"])
        sheet.row_dimensions[row_index].height = 96
        # 找第一张能解析到本地的图
        for image in images:
            local_path = _resolve_local_image(image, data_root)
            if local_path is None:
                continue
            img_bytes = _maybe_resize_cover_image(local_path)
            cover_cell = sheet.cell(row=row_index, column=9)
            try:
                xlsx_image = XlsxImage(BytesIO(img_bytes))
                sheet.add_image(xlsx_image, cover_cell.coordinate)
            except Exception as exc:  # pragma: no cover - 解码失败
                logger.warning("封面图嵌入失败 note_id=%s path=%s: %s", note.id, local_path, exc)
            break
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
